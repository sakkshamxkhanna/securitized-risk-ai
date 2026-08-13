"""Excel workbook export for the surveillance pack.

Builds the workbook a desk analyst would circulate: a loan tape, live
stratification tables, cashflow and waterfall schedules, exposure/RWA,
and scenario stress.

Design principle: the stratification and exposure tables are computed by
**Excel formulas over the loan tape**, not pasted values. Change a
severity or LGD assumption on the Assumptions sheet and the workbook
recalculates. A workbook of hardcoded numbers is a screenshot; this one
is a model.

Run: python -m src.excel_export
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

FONT = "Arial"
BLUE = "0000FF"      # hardcoded inputs
BLACK = "000000"     # formulas
GREEN = "008000"     # cross-sheet links
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);-'
CUR2 = '$#,##0.00;($#,##0.00);-'
PCT1 = '0.0%;(0.0%);-'
PCT2 = '0.00%;(0.00%);-'
NUM = '#,##0;(#,##0);-'


def _title(ws, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=13, bold=True, color="1F3864")


def _header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _style_data(ws, first_row: int, last_row: int, first_col: int, last_col: int,
                 band: bool = True) -> None:
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if band and (r - first_row) % 2 == 1:
                cell.fill = BAND_FILL


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1,
               number_formats: dict[str, str] | None = None) -> int:
    """Writes a dataframe as a styled block. Returns the last data row."""
    _header_row(ws, start_row, list(df.columns), start_col)
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = row[col]
            if hasattr(val, "item"):
                val = val.item()
            c = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            if number_formats and col in number_formats:
                c.number_format = number_formats[col]
    last = start_row + len(df)
    _style_data(ws, start_row + 1, last, start_col, start_col + len(df.columns) - 1)
    return last


# ----------------------------------------------------------------- sheets

def _assumptions_sheet(wb: Workbook, metrics: dict) -> None:
    ws = wb.create_sheet("Assumptions")
    _title(ws, "A1", "Surveillance Assumptions & Model Inputs")
    ws["A2"] = "Blue cells are hardcoded inputs. Every downstream sheet references them."
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    rows = [
        ("Deal", "Synthetic RMBS Pool 2026-1", None, "Demonstration pool — synthetic data"),
        ("Projection horizon (months)", 36, NUM, "Surveillance projection window"),
        ("Loss severity (LGD)", 0.35, PCT1, "Applied to defaulted balance"),
        ("Escalation threshold (% UPB)", 0.02, PCT1, "Cumulative loss above this escalates to desk"),
        ("Senior tranche size", 0.78, PCT1, "Capital structure — Class A"),
        ("Mezzanine tranche size", 0.14, PCT1, "Capital structure — Class M"),
        ("Subordinate tranche size", 0.08, PCT1, "Capital structure — Class B"),
    ]
    _header_row(ws, 4, ["Assumption", "Value", "", "Note"])
    for i, (label, val, fmt, note) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
        c = ws.cell(row=r, column=2, value=val)
        c.font = Font(name=FONT, size=10, bold=True, color=BLUE)
        c.fill = INPUT_FILL
        if fmt:
            c.number_format = fmt
        ws.cell(row=r, column=4, value=note).font = Font(name=FONT, size=9, color="595959")
    _style_data(ws, 5, 4 + len(rows), 1, 4, band=False)

    # Risk weight lookup table, referenced by the Exposure sheet via INDEX/MATCH.
    _title(ws, "A15", "Standardized-approach risk weights")
    ws["A16"] = "Simplified rating-band proxy — not a full SEC-SA / SEC-IRBA implementation."
    ws["A16"].font = Font(name=FONT, size=9, italic=True, color="595959")
    _header_row(ws, 17, ["Rating", "Risk weight"])
    for i, (rating, rw) in enumerate(
            [("AAA", 0.20), ("AA", 0.20), ("A", 0.50), ("BBB", 1.00),
             ("BB", 2.50), ("B", 4.25), ("Unrated", 12.50)]):
        r = 18 + i
        ws.cell(row=r, column=1, value=rating).font = Font(name=FONT, size=10)
        c = ws.cell(row=r, column=2, value=rw)
        c.font = Font(name=FONT, size=10, color=BLUE)
        c.fill = INPUT_FILL
        c.number_format = '0.00"x"'
    _style_data(ws, 18, 24, 1, 2, band=False)

    v = metrics.get("forecaster_validation", {})
    _title(ws, "A27", "Model validation (holdout, 400 unseen sequences)")
    _header_row(ws, 28, ["Metric", "CPR", "CDR"])
    for i, (label, a, b, fmt) in enumerate([
            ("R²", v.get("cpr_r2"), v.get("cdr_r2"), '0.000'),
            ("MAE", v.get("cpr_mae"), v.get("cdr_mae"), '0.00000'),
            ("Predicted sd", None, v.get("cdr_pred_sd"), '0.00000'),
            ("Realised sd", None, v.get("cdr_true_sd"), '0.00000')]):
        r = 29 + i
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
        for j, val in enumerate((a, b)):
            if val is not None:
                c = ws.cell(row=r, column=2 + j, value=val)
                c.number_format = fmt
    _style_data(ws, 29, 32, 1, 3, band=False)

    _autosize(ws, {"A": 30, "B": 16, "C": 14, "D": 52})


def _loan_tape_sheet(wb: Workbook, pool: pd.DataFrame) -> None:
    ws = wb.create_sheet("Loan Tape")
    _title(ws, "A1", "Loan-Level Collateral Tape")
    ws["A2"] = f"{len(pool):,} loans. Source data for every stratification on this workbook."
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    cols = ["loan_id", "orig_balance", "current_balance", "coupon", "fico", "ltv", "dti",
            "seasoning_mo", "remaining_term_mo", "state", "property_type", "originator", "servicer"]
    df = pool[cols]
    _write_df(ws, df, start_row=4, number_formats={
        "orig_balance": CUR, "current_balance": CUR, "coupon": '0.000"%"',
        "ltv": '0.0', "dti": '0.0', "fico": NUM,
        "seasoning_mo": NUM, "remaining_term_mo": NUM})

    last = 4 + len(df)
    ref = f"A4:{get_column_letter(len(cols))}{last}"
    table = Table(displayName="LoanTape", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A5"
    _autosize(ws, {"A": 12, "B": 15, "C": 16, "D": 10, "E": 8, "F": 8, "G": 8,
                    "H": 12, "I": 14, "J": 8, "K": 14, "L": 22, "M": 22})


def _stratification_sheet(wb: Workbook, pool: pd.DataFrame) -> None:
    """Stratifications computed by formula over the loan tape, so edits to
    the tape flow through — this is the tape-cutting a desk actually does."""
    ws = wb.create_sheet("Stratification")
    n = len(pool)
    tape_lo, tape_hi = 5, 4 + n
    bal = f"'Loan Tape'!$C${tape_lo}:$C${tape_hi}"
    fico = f"'Loan Tape'!$E${tape_lo}:$E${tape_hi}"
    ltv = f"'Loan Tape'!$F${tape_lo}:$F${tape_hi}"
    cpn = f"'Loan Tape'!$D${tape_lo}:$D${tape_hi}"
    state = f"'Loan Tape'!$J${tape_lo}:$J${tape_hi}"

    _title(ws, "A1", "Collateral Stratification")
    ws["A2"] = ("Every figure below is a live formula over the Loan Tape "
                "(SUMIFS / SUMPRODUCT). Edit the tape and these update.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    # ---- FICO
    _title(ws, "A4", "By FICO")
    headers = ["Bucket", "Loan count", "UPB", "% of pool", "WA coupon", "WA LTV"]
    _header_row(ws, 5, headers)
    fico_bands = [("<620", 0, 620), ("620-659", 620, 660), ("660-699", 660, 700),
                  ("700-739", 700, 740), ("740-779", 740, 780), ("780+", 780, 10000)]
    r0 = 6
    for i, (label, lo, hi) in enumerate(fico_bands):
        r = r0 + i
        crit = f'">={lo}",{fico},"<{hi}"'
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f'=COUNTIFS({fico},{crit})').number_format = NUM
        ws.cell(row=r, column=3, value=f'=SUMIFS({bal},{fico},{crit})').number_format = CUR
        ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/$C${r0 + len(fico_bands)},0)').number_format = PCT2
        ws.cell(row=r, column=5, value=(
            f'=IFERROR(SUMPRODUCT(({fico}>={lo})*({fico}<{hi})*{bal}*{cpn})/C{r},0)')
        ).number_format = '0.000"%"'
        ws.cell(row=r, column=6, value=(
            f'=IFERROR(SUMPRODUCT(({fico}>={lo})*({fico}<{hi})*{bal}*{ltv})/C{r},0)')
        ).number_format = '0.0'
    tot = r0 + len(fico_bands)
    ws.cell(row=tot, column=1, value="Total")
    ws.cell(row=tot, column=2, value=f'=SUM(B{r0}:B{tot-1})').number_format = NUM
    ws.cell(row=tot, column=3, value=f'=SUM(C{r0}:C{tot-1})').number_format = CUR
    ws.cell(row=tot, column=4, value=f'=SUM(D{r0}:D{tot-1})').number_format = PCT2
    for c in range(1, 7):
        ws.cell(row=tot, column=c).font = Font(name=FONT, size=10, bold=True)
    _style_data(ws, r0, tot, 1, 6)

    # ---- LTV
    ltv_start = tot + 3
    _title(ws, f"A{ltv_start}", "By LTV")
    _header_row(ws, ltv_start + 1, headers)
    ltv_bands = [("<60", 0, 60), ("60-69", 60, 70), ("70-79", 70, 80),
                 ("80-89", 80, 90), ("90-99", 90, 100), ("100+", 100, 1000)]
    r0 = ltv_start + 2
    for i, (label, lo, hi) in enumerate(ltv_bands):
        r = r0 + i
        crit = f'">={lo}",{ltv},"<{hi}"'
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f'=COUNTIFS({ltv},{crit})').number_format = NUM
        ws.cell(row=r, column=3, value=f'=SUMIFS({bal},{ltv},{crit})').number_format = CUR
        ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/$C${r0 + len(ltv_bands)},0)').number_format = PCT2
        ws.cell(row=r, column=5, value=(
            f'=IFERROR(SUMPRODUCT(({ltv}>={lo})*({ltv}<{hi})*{bal}*{cpn})/C{r},0)')
        ).number_format = '0.000"%"'
        ws.cell(row=r, column=6, value=(
            f'=IFERROR(SUMPRODUCT(({ltv}>={lo})*({ltv}<{hi})*{bal}*{fico})/C{r},0)')
        ).number_format = NUM
    ws.cell(row=ltv_start + 1, column=6, value="WA FICO")
    ws.cell(row=ltv_start + 1, column=6).font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    ws.cell(row=ltv_start + 1, column=6).fill = HEADER_FILL
    tot2 = r0 + len(ltv_bands)
    ws.cell(row=tot2, column=1, value="Total")
    ws.cell(row=tot2, column=2, value=f'=SUM(B{r0}:B{tot2-1})').number_format = NUM
    ws.cell(row=tot2, column=3, value=f'=SUM(C{r0}:C{tot2-1})').number_format = CUR
    ws.cell(row=tot2, column=4, value=f'=SUM(D{r0}:D{tot2-1})').number_format = PCT2
    for c in range(1, 7):
        ws.cell(row=tot2, column=c).font = Font(name=FONT, size=10, bold=True)
    _style_data(ws, r0, tot2, 1, 6)

    # ---- Geography
    geo_start = tot2 + 3
    _title(ws, f"A{geo_start}", "By state")
    _header_row(ws, geo_start + 1, ["State", "Loan count", "UPB", "% of pool", "WA FICO"])
    states = sorted(pool["state"].unique())
    r0 = geo_start + 2
    for i, st in enumerate(states):
        r = r0 + i
        ws.cell(row=r, column=1, value=st)
        ws.cell(row=r, column=2, value=f'=COUNTIFS({state},A{r})').number_format = NUM
        ws.cell(row=r, column=3, value=f'=SUMIFS({bal},{state},A{r})').number_format = CUR
        ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/$C${r0 + len(states)},0)').number_format = PCT2
        ws.cell(row=r, column=5, value=(
            f'=IFERROR(SUMPRODUCT(({state}=A{r})*{bal}*{fico})/C{r},0)')).number_format = NUM
    tot3 = r0 + len(states)
    ws.cell(row=tot3, column=1, value="Total")
    ws.cell(row=tot3, column=2, value=f'=SUM(B{r0}:B{tot3-1})').number_format = NUM
    ws.cell(row=tot3, column=3, value=f'=SUM(C{r0}:C{tot3-1})').number_format = CUR
    ws.cell(row=tot3, column=4, value=f'=SUM(D{r0}:D{tot3-1})').number_format = PCT2
    for c in range(1, 6):
        ws.cell(row=tot3, column=c).font = Font(name=FONT, size=10, bold=True)
    _style_data(ws, r0, tot3, 1, 5)

    # concentration heat: largest state exposures stand out
    ws.conditional_formatting.add(
        f"D{r0}:D{tot3-1}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="F8CBAD"))

    _autosize(ws, {"A": 16, "B": 13, "C": 18, "D": 12, "E": 13, "F": 12})


def _cashflow_sheet(wb: Workbook, cashflows: pd.DataFrame) -> None:
    ws = wb.create_sheet("Pool Cashflows")
    _title(ws, "A1", "Projected Pool Cashflows — Base Case")
    ws["A2"] = ("Model-projected. CPR/CDR from the Transformer forecaster, "
                "default losses at the severity on Assumptions.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    df = cashflows.copy()
    last = _write_df(ws, df, start_row=4, number_formats={
        "scheduled_principal": CUR, "prepayment": CUR, "interest_collected": CUR,
        "loss": CUR, "ending_balance": CUR, "period": NUM})

    r = last + 1
    ws.cell(row=r, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    for col in range(2, 6):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f'=SUM({L}5:{L}{last})')
        c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = CUR
    ws.freeze_panes = "A5"
    _autosize(ws, {"A": 10, "B": 20, "C": 16, "D": 18, "E": 14, "F": 18})


def _waterfall_sheet(wb: Workbook, schedule: pd.DataFrame, summary: pd.DataFrame) -> None:
    ws = wb.create_sheet("Waterfall")
    _title(ws, "A1", "Tranche Waterfall — Sequential Pay")
    ws["A2"] = ("Interest pro-rata to outstanding balance; principal sequential senior→sub; "
                "losses reverse-sequential (subordinate absorbs first).")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    _title(ws, "A4", "Tranche summary")
    last = _write_df(ws, summary, start_row=5, number_formats={
        "total_principal": CUR, "total_interest": CUR, "final_balance": CUR})

    sched_start = last + 3
    _title(ws, f"A{sched_start}", "Period detail")
    _write_df(ws, schedule.drop(columns=["beginning_balance_note"], errors="ignore"),
              start_row=sched_start + 1, number_formats={
                  "principal_paid": CUR, "interest_paid": CUR,
                  "ending_balance": CUR, "period": NUM})
    _autosize(ws, {"A": 10, "B": 20, "C": 18, "D": 18, "E": 18})


def _exposure_sheet(wb: Workbook, exposure: pd.DataFrame) -> None:
    """EL and RWA are formulas referencing the Assumptions sheet, so changing
    LGD or a risk weight there reprices the whole table."""
    ws = wb.create_sheet("Exposure & RWA")
    _title(ws, "A1", "Tranche Exposure, Expected Loss and RWA")
    ws["A2"] = ("EL = EAD × PD × LGD and RWA = EAD × risk weight are live formulas. "
                "LGD and the risk-weight table live on Assumptions.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    _header_row(ws, 4, ["Tranche", "Proxy rating", "EAD", "PD", "LGD",
                         "Expected loss", "Risk weight", "RWA"])
    r0 = 5
    for i, (_, row) in enumerate(exposure.iterrows()):
        r = r0 + i
        ws.cell(row=r, column=1, value=row["tranche"])
        ws.cell(row=r, column=2, value=row["proxy_rating"])
        ws.cell(row=r, column=3, value=float(row["ead"])).number_format = CUR
        ws.cell(row=r, column=4, value=float(row["pd"])).number_format = PCT2
        # LGD pulled from Assumptions — green marks a cross-sheet link.
        c = ws.cell(row=r, column=5, value="=Assumptions!$B$7")
        c.number_format = PCT1
        c.font = Font(name=FONT, size=10, color=GREEN)
        ws.cell(row=r, column=6, value=f"=C{r}*D{r}*E{r}").number_format = CUR
        c = ws.cell(row=r, column=7,
                    value=f'=INDEX(Assumptions!$B$18:$B$24,MATCH(B{r},Assumptions!$A$18:$A$24,0))')
        c.number_format = '0.00"x"'
        c.font = Font(name=FONT, size=10, color=GREEN)
        ws.cell(row=r, column=8, value=f"=C{r}*G{r}").number_format = CUR

    tot = r0 + len(exposure)
    ws.cell(row=tot, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    for col, fmt in ((3, CUR), (6, CUR), (8, CUR)):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f'=SUM({L}{r0}:{L}{tot-1})')
        c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = fmt
    _style_data(ws, r0, tot, 1, 8)
    _autosize(ws, {"A": 20, "B": 14, "C": 18, "D": 10, "E": 10,
                    "F": 16, "G": 12, "H": 18})


def _scenario_sheet(wb: Workbook, stress: pd.DataFrame, upb: float) -> None:
    ws = wb.create_sheet("Scenario Stress")
    _title(ws, "A1", "VAE-Generated Scenario Stress")
    ws["A2"] = ("Scenarios sampled from the latent space of a VAE trained on macro regime "
                "paths — generated, not hand-specified.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    _header_row(ws, 4, ["Scenario", "Label", "Avg CPR", "Avg CDR", "Cumulative loss",
                         "Loss % of UPB", "Sub write-down", "Escalates?"])
    r0 = 5
    for i, (_, row) in enumerate(stress.iterrows()):
        r = r0 + i
        ws.cell(row=r, column=1, value=row["scenario"])
        ws.cell(row=r, column=2, value=row["label"])
        ws.cell(row=r, column=3, value=float(row["avg_cpr_pct"]) / 100).number_format = PCT2
        ws.cell(row=r, column=4, value=float(row["avg_cdr_pct"]) / 100).number_format = PCT2
        ws.cell(row=r, column=5, value=float(row["cumulative_loss"])).number_format = CUR
        ws.cell(row=r, column=6, value=f"=E{r}/$B$1").number_format = PCT2
        ws.cell(row=r, column=7, value=float(row["sub_tranche_writedown"])).number_format = CUR
        ws.cell(row=r, column=8,
                value=f'=IF(F{r}>Assumptions!$B$8,"ESCALATE","within tolerance")')

    # pool UPB parked in a referenced cell so the % column is a real formula
    ws["B1"] = upb
    ws["B1"].number_format = CUR
    ws["B1"].font = Font(name=FONT, size=9, color=BLUE)
    ws["C1"] = "← pool UPB (denominator for loss %)"
    ws["C1"].font = Font(name=FONT, size=9, italic=True, color="595959")

    last = r0 + len(stress) - 1
    _style_data(ws, r0, last, 1, 8)
    ws.conditional_formatting.add(
        f"H{r0}:H{last}",
        CellIsRule(operator="equal", formula=['"ESCALATE"'],
                   fill=PatternFill("solid", fgColor="FFC7CE"),
                   font=Font(name=FONT, size=10, bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        f"F{r0}:F{last}",
        ColorScaleRule(start_type="min", start_color="C6EFCE",
                       end_type="max", end_color="FFC7CE"))
    _autosize(ws, {"A": 11, "B": 44, "C": 11, "D": 11, "E": 18, "F": 14,
                    "G": 18, "H": 18})


def _summary_sheet(wb: Workbook, metrics: dict) -> None:
    ws = wb.create_sheet("Summary", 0)
    _title(ws, "A1", "Securitized Products Surveillance Pack")
    ws["A2"] = f"Synthetic RMBS Pool 2026-1 · generated {metrics.get('generated_at', 'n/a')[:10]}"
    ws["A2"].font = Font(name=FONT, size=10, color="595959")
    ws["A3"] = ("All data is synthetic and generated for demonstration. "
                "RWA is a simplified rating-band proxy, not SEC-SA / SEC-IRBA.")
    ws["A3"].font = Font(name=FONT, size=9, italic=True, color="C00000")

    s = metrics["pool_summary"]
    _header_row(ws, 5, ["Pool metric", "Value"])
    pool_rows = [
        ("Loan count", s["loan_count"], NUM),
        ("Current UPB", s["total_upb"], CUR),
        ("WA coupon (%)", s["wa_coupon"], '0.000'),
        ("WA FICO", s["wa_fico"], NUM),
        ("WA LTV (%)", s["wa_ltv"], '0.0'),
        ("WA DTI (%)", s["wa_dti"], '0.0'),
        ("WA seasoning (mo)", s["wa_seasoning_mo"], '0.0'),
    ]
    for i, (label, val, fmt) in enumerate(pool_rows):
        r = 6 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val).number_format = fmt
    _style_data(ws, 6, 5 + len(pool_rows), 1, 2)

    _header_row(ws, 15, ["Projected performance", "Value"])
    perf = [
        ("Average CPR", metrics["avg_cpr"], PCT2),
        ("Average CDR (GNN-adjusted)", metrics["avg_cdr"], PCT2),
        ("GNN risk multiplier", metrics["gnn_avg_multiplier"], '0.000"x"'),
        ("Cumulative loss (base)", metrics["cumulative_loss_base"], CUR),
        ("Total RWA", metrics.get("total_rwa"), CUR),
        ("Total expected loss", metrics.get("total_expected_loss"), CUR),
    ]
    for i, (label, val, fmt) in enumerate(perf):
        r = 16 + i
        ws.cell(row=r, column=1, value=label)
        if val is not None:
            ws.cell(row=r, column=2, value=val).number_format = fmt
    _style_data(ws, 16, 15 + len(perf), 1, 2)

    esc = metrics.get("escalations", [])
    _title(ws, "A24", "Escalations")
    if esc:
        for i, e in enumerate(esc):
            c = ws.cell(row=25 + i, column=1, value=e)
            c.font = Font(name=FONT, size=10, bold=True, color="9C0006")
            c.fill = PatternFill("solid", fgColor="FFC7CE")
    else:
        ws["A25"] = "None — all scenarios within tolerance."
        ws["A25"].font = Font(name=FONT, size=10, color="006100")

    _title(ws, "A29", "Contents")
    contents = [
        ("Assumptions", "Inputs, risk weights, model validation"),
        ("Loan Tape", "Loan-level collateral data"),
        ("Stratification", "FICO / LTV / geography cuts (live formulas)"),
        ("Pool Cashflows", "Projected monthly cashflows, base case"),
        ("Waterfall", "Tranche principal, interest, balances"),
        ("Exposure & RWA", "EAD, PD, LGD, EL, RWA"),
        ("Scenario Stress", "Generated scenarios and escalation flags"),
    ]
    for i, (sheet, desc) in enumerate(contents):
        r = 30 + i
        ws.cell(row=r, column=1, value=sheet).font = Font(name=FONT, size=10, bold=True)
        ws.cell(row=r, column=2, value=desc).font = Font(name=FONT, size=10)

    _autosize(ws, {"A": 34, "B": 46})


def build_workbook(output_path: Path | None = None) -> Path:
    out = output_path or (OUTPUT_DIR / "surveillance_pack.xlsx")
    metrics = json.loads((OUTPUT_DIR / "run_metrics.json").read_text())

    pool = pd.read_csv(OUTPUT_DIR / "loan_tape.csv")
    cashflows = pd.read_csv(OUTPUT_DIR / "pool_cashflows_base.csv")
    schedule = pd.read_csv(OUTPUT_DIR / "waterfall_schedule_base.csv")
    exposure = pd.read_csv(OUTPUT_DIR / "exposure_rwa.csv")
    stress = pd.read_csv(OUTPUT_DIR / "scenario_stress.csv")

    summary = schedule.groupby("tranche", as_index=False).agg(
        total_principal=("principal_paid", "sum"),
        total_interest=("interest_paid", "sum"),
        final_balance=("ending_balance", "last"))

    wb = Workbook()
    wb.remove(wb.active)
    _summary_sheet(wb, metrics)
    _assumptions_sheet(wb, metrics)
    _loan_tape_sheet(wb, pool)
    _stratification_sheet(wb, pool)
    _cashflow_sheet(wb, cashflows)
    _waterfall_sheet(wb, schedule, summary)
    _exposure_sheet(wb, exposure)
    _scenario_sheet(wb, stress, float(metrics["pool_summary"]["total_upb"]))

    wb.save(out)
    return out


if __name__ == "__main__":
    path = build_workbook()
    print(f"wrote {path}")
